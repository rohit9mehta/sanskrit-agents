"""Phase 2, item 3b: score grader responses.

Reads data/eval/packets/responses/grader_*.csv (label answers A/B/C/tie/-),
maps labels to systems via the answer key, and reports per criterion:
win counts, agent-vs-each-baseline exact sign tests (ties dropped),
inter-grader agreement (Cohen's kappa), and the hard-subset cut.

Stdlib-only. Run after graders return their CSVs.
"""

import csv
import json
import math
import sys
from collections import Counter
from itertools import combinations
from pathlib import Path

from shastrartha.texts import DATA_DIR

EVAL = DATA_DIR / "eval"
RESPONSES = EVAL / "packets" / "responses"
CRITERIA = ["term_fidelity", "compound_resolution", "overall"]


def binom_two_sided(k: int, n: int) -> float:
    """Exact two-sided sign test p-value, p0 = 0.5."""
    if n == 0:
        return 1.0
    p_k = math.comb(n, k) / 2**n
    return min(1.0, sum(math.comb(n, i) / 2**n
                        for i in range(n + 1)
                        if math.comb(n, i) / 2**n <= p_k + 1e-12))


def kappa(a: list[str], b: list[str]) -> float:
    assert len(a) == len(b)
    n = len(a)
    if n == 0:
        return float("nan")
    po = sum(x == y for x, y in zip(a, b)) / n
    ca, cb = Counter(a), Counter(b)
    pe = sum(ca[k] * cb[k] for k in set(ca) | set(cb)) / n**2
    return (po - pe) / (1 - pe) if pe < 1 else 1.0


# Excel answer sheets (scripts/16) use friendly headers; map them back.
XLSX_HEADER_MAP = {
    "Verse": "verse",
    "Q1: Technical terms (A/B/C/tie)": "term_fidelity",
    "Q2: Compounds (A/B/C/tie)": "compound_resolution",
    "Q3: Overall (A/B/C/tie)": "overall",
    "Comments (optional)": "comments",
}


def _rows_from_file(path: Path) -> list[dict]:
    if path.suffix.lower() == ".xlsx":
        from openpyxl import load_workbook

        ws = load_workbook(path, read_only=True).active
        rows = list(ws.iter_rows(values_only=True))
        header_i = next(i for i, r in enumerate(rows)
                        if r and str(r[0]).strip().lower() == "verse")
        headers = [XLSX_HEADER_MAP.get(str(h).strip(), str(h).strip())
                   for h in rows[header_i]]
        return [
            {h: ("" if v is None else str(v)) for h, v in zip(headers, r)}
            for r in rows[header_i + 1:] if r and r[0] is not None
        ]
    with path.open() as f:
        return list(csv.DictReader(f))


def load_grader(path: Path, key: dict) -> dict:
    out = {}
    for row in _rows_from_file(path):
        n = str(row.get("verse", "")).strip()
        if n.endswith(".0"):  # Excel numeric cells round-trip as floats
            n = n[:-2]
        if not n.isdigit():
            continue
        ans = {}
        for c in CRITERIA:
            v = (row.get(c) or "").strip().upper()
            ans[c] = key[n].get(v, v.lower()) if v in ("A", "B", "C") else (v.lower() or None)
        out[int(n)] = ans
    return out


def main() -> int:
    key = json.loads((EVAL / "answer_key.json").read_text())["labels"]
    hard = set(json.loads((EVAL / "answer_key.json").read_text())["hard_subset"])
    files = (sorted(RESPONSES.glob("grader_*.csv")) + sorted(RESPONSES.glob("grader_*.xlsx"))
             if RESPONSES.exists() else [])
    files = [f for f in files if "TEMPLATE" not in f.name]
    if not files:
        print("no grader responses in", RESPONSES, "— nothing to score yet")
        return 0
    graders = {f.stem: load_grader(f, key) for f in files}
    print(f"graders: {list(graders)}\n")

    for subset_name, verses in [("all 30", set(range(1, 31))), ("hard subset", hard)]:
        print(f"== {subset_name} ==")
        for c in CRITERIA:
            wins = Counter()
            pair = {"agent>raw_llm": 0, "raw_llm>agent": 0,
                    "agent>mitra": 0, "mitra>agent": 0}
            for g in graders.values():
                for n, ans in g.items():
                    if n not in verses or not ans.get(c):
                        continue
                    w = ans[c]
                    wins[w] += 1
                    if w == "agent":
                        pair["agent>raw_llm"] += 1
                        pair["agent>mitra"] += 1
                    elif w == "raw_llm":
                        pair["raw_llm>agent"] += 1
                    elif w == "mitra":
                        pair["mitra>agent"] += 1
            p_raw = binom_two_sided(pair["agent>raw_llm"],
                                    pair["agent>raw_llm"] + pair["raw_llm>agent"])
            p_mit = binom_two_sided(pair["agent>mitra"],
                                    pair["agent>mitra"] + pair["mitra>agent"])
            print(f"  {c}: wins {dict(wins)} | agent vs raw p={p_raw:.3f} "
                  f"| agent vs mitra p={p_mit:.3f}")
        for (na, ga), (nb, gb) in combinations(graders.items(), 2):
            for c in CRITERIA:
                common = [n for n in sorted(set(ga) & set(gb) & verses)
                          if ga[n].get(c) and gb[n].get(c)]
                k = kappa([ga[n][c] for n in common], [gb[n][c] for n in common])
                print(f"  kappa({na},{nb}) {c}: {k:.2f} (n={len(common)})")
        print()
    return 0


if __name__ == "__main__":
    sys.exit(main())
