"""Render the grader materials in scholar-friendly formats:

  eval_packet.pdf         — typeset from eval_packet.md (parsed, so A/B/C
                            order is guaranteed identical to the answer key)
  Response_Sheet_1.xlsx   — answer sheet with A/B/C/tie dropdowns
  Response_Sheet_2.xlsx   — identical copy for the second grader

Fonts: Times New Roman (macOS Supplemental) with Arial Unicode fallback —
both cover Latin Extended Additional (ā ī ū ṛ ṃ ḥ ś ṣ ñ ṅ ṇ ṭ ḍ).
"""

import re
import sys

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle
from reportlab.lib.units import cm
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.platypus import (HRFlowable, KeepTogether, Paragraph,
                                SimpleDocTemplate, Spacer)

from shastrartha.texts import DATA_DIR

PACKETS = DATA_DIR / "eval" / "packets"
MD = PACKETS / "eval_packet.md"
PDF = PACKETS / "eval_packet.pdf"

FONT_DIR = "/System/Library/Fonts/Supplemental"


def register_fonts() -> None:
    pdfmetrics.registerFont(TTFont("TNR", f"{FONT_DIR}/Times New Roman.ttf"))
    pdfmetrics.registerFont(TTFont("TNR-Bold", f"{FONT_DIR}/Times New Roman Bold.ttf"))
    pdfmetrics.registerFont(TTFont("TNR-Italic", f"{FONT_DIR}/Times New Roman Italic.ttf"))


S = {}


def styles() -> None:
    S["title"] = ParagraphStyle("title", fontName="TNR-Bold", fontSize=17, leading=22,
                                spaceAfter=10)
    S["body"] = ParagraphStyle("body", fontName="TNR", fontSize=11, leading=15.5,
                               spaceAfter=6)
    S["vhead"] = ParagraphStyle("vhead", fontName="TNR-Bold", fontSize=13, leading=17,
                                spaceBefore=6, spaceAfter=6)
    S["sanskrit"] = ParagraphStyle("sanskrit", fontName="TNR", fontSize=11.5,
                                   leading=16, leftIndent=0.4 * cm, spaceAfter=6)
    S["sys"] = ParagraphStyle("sys", fontName="TNR", fontSize=11, leading=15.5,
                              leftIndent=0.4 * cm, spaceAfter=5)


def md_inline(text: str) -> str:
    text = text.replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")
    text = re.sub(r"\*\*(.+?)\*\*", r"<b>\1</b>", text)
    text = re.sub(r"(?<!\*)\*([^*]+)\*(?!\*)", r"<i>\1</i>", text)
    return text


def build_pdf() -> None:
    raw = MD.read_text(encoding="utf-8")
    intro, *verses = re.split(r"\n## Verse ", raw)

    story = []
    intro = intro.replace("\n---\n", "\n").strip()
    first, *rest = intro.split("\n\n")
    story.append(Paragraph(md_inline(first.lstrip("# ").strip()), S["title"]))
    for para in rest:
        para = para.strip()
        if not para:
            continue
        if re.match(r"^\d+\.", para):
            for item in re.split(r"\n(?=\d+\.)", para):
                story.append(Paragraph(md_inline(item.replace("\n", " ")), S["body"]))
        else:
            story.append(Paragraph(md_inline(para.replace("\n", " ")), S["body"]))
    story.append(Spacer(1, 8))
    story.append(HRFlowable(width="100%", thickness=0.7))

    for block in verses:
        n, _, rest = block.partition("\n")
        chunk = [Paragraph(f"Verse {n.strip()}", S["vhead"])]
        # Paragraph-wise, not line-wise: translations contain hard newlines
        # (verse-shaped output) which stay inside their indented block.
        for para in rest.split("\n\n"):
            para = para.strip()
            if not para or para == "---":
                continue
            style = S["sanskrit"] if para.startswith("**Sanskrit") else (
                S["sys"] if re.match(r"^\*\*[ABC]\.\*\*", para) else S["body"])
            chunk.append(Paragraph(md_inline(para).replace("\n", "<br/>"), style))
        story.append(KeepTogether(chunk))
        story.append(Spacer(1, 4))
        story.append(HRFlowable(width="100%", thickness=0.4, color="#999999"))
        story.append(Spacer(1, 4))

    doc = SimpleDocTemplate(str(PDF), pagesize=A4,
                            leftMargin=2.2 * cm, rightMargin=2.2 * cm,
                            topMargin=2 * cm, bottomMargin=2 * cm,
                            title="Trimsika translation evaluation")
    doc.build(story)
    print(f"wrote {PDF}")


HEADERS = [
    "Verse",
    "Q1: Technical terms (A/B/C/tie)",
    "Q2: Compounds (A/B/C/tie)",
    "Q3: Overall (A/B/C/tie)",
    "Comments (optional)",
]


def build_sheet(path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "responses"

    ws.merge_cells("A1:E1")
    intro = ws["A1"]
    intro.value = ("For each verse (see the PDF packet), answer the three questions "
                   "by choosing A, B, C, or tie in each column. Leave blank only if "
                   "no option is acceptable, and please say why in Comments.")
    intro.font = Font(italic=True)
    intro.alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[1].height = 44

    ws.append(HEADERS)
    for c in range(1, 6):
        cell = ws.cell(row=2, column=c)
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="DDDDDD")
        cell.alignment = Alignment(wrap_text=True, vertical="top")
    for n in range(1, 31):
        ws.append([n, "", "", "", ""])

    dv = DataValidation(type="list", formula1='"A,B,C,tie"', allow_blank=True,
                        showInputMessage=True, showErrorMessage=True)
    dv.promptTitle = "Your judgment"
    dv.prompt = "Choose A, B, C, or tie"
    dv.errorTitle = "Invalid entry"
    dv.error = "Please enter A, B, C, or tie"
    ws.add_data_validation(dv)
    dv.add("B3:D32")

    for col, width in zip("ABCDE", (7, 15, 15, 15, 58)):
        ws.column_dimensions[col].width = width
    for row in ws.iter_rows(min_row=3, max_row=32, min_col=5, max_col=5):
        row[0].alignment = Alignment(wrap_text=True)
    ws.freeze_panes = "A3"
    wb.save(path)
    print(f"wrote {path}")


def main() -> int:
    register_fonts()
    styles()
    build_pdf()
    build_sheet(PACKETS / "Response_Sheet_1.xlsx")
    build_sheet(PACKETS / "Response_Sheet_2.xlsx")
    return 0


if __name__ == "__main__":
    sys.exit(main())
